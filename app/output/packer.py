"""Build snapshot ZIP and XLSX from the live append-only data files.

Both run while the harvester is still writing. We open files for read in
text mode; appends from the harvester are atomic line writes, so a snapshot
contains all complete records up to the moment the read is performed.
"""
from __future__ import annotations

import json
import time
import zipfile
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook

from app.core.logger import get_logger

log = get_logger("packer")


def _iter_jsonl(path: Path):
    if not path.exists():
        return
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                yield json.loads(line)
            except Exception:
                continue


def build_snapshot(data_dir: Path, out_dir: Path, sources: Dict[str, List[str]]) -> tuple[Path, Path]:
    """Build zip + xlsx from current data. Returns (zip_path, xlsx_path).

    Only sources that actually contain data are included in the XLSX
    (Excel chokes at hundreds of empty sheets). The ZIP includes the
    raw .txt for every non-empty source.
    """
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = time.strftime("%Y%m%d-%H%M%S")
    zip_path = out_dir / f"snapshot-{ts}.zip"
    xlsx_path = out_dir / f"snapshot-{ts}.xlsx"

    nonempty: list[str] = []
    for src in sources:
        p = data_dir / f"{src}.txt"
        if p.exists() and p.stat().st_size > 0:
            nonempty.append(src)

    # ZIP: include each non-empty source.txt
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=6) as zf:
        for src in nonempty:
            zf.write(data_dir / f"{src}.txt", arcname=f"{src}.txt")
        readme = (
            "parser-combine snapshot\n"
            f"taken_at: {ts}\n"
            f"total_sources_registered: {len(sources)}\n"
            f"sources_with_data: {len(nonempty)}\n"
            f"sources: {', '.join(nonempty)}\n"
        )
        zf.writestr("README.txt", readme)

    # XLSX: one sheet per source that has data. To avoid Excel sheet-name
    # collisions when names are truncated to 31 chars, deduplicate names.
    wb = Workbook(write_only=True)
    used: set[str] = set()
    if not nonempty:
        wb.create_sheet(title="empty")
    else:
        for src in nonempty:
            cols = sources[src]
            base = src[:28] or "sheet"
            title = base
            i = 2
            while title.lower() in used:
                title = f"{base[:25]}_{i}"
                i += 1
            used.add(title.lower())
            ws = wb.create_sheet(title=title[:31])
            ws.append(list(cols))
            for row in _iter_jsonl(data_dir / f"{src}.jsonl"):
                ws.append([row.get(c, "") for c in cols])
    wb.save(xlsx_path)
    return zip_path, xlsx_path
